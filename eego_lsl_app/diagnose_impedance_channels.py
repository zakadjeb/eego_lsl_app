"""Systematic impedance-channel diagnostic for ANT/eego SDK.

This script is intentionally diagnostic rather than pretty. It records *exactly*
what the SDK returns for an impedance stream/state:

  - amplifier.getChannelList() before opening impedance
  - impedanceStream.getChannelList() after opening impedance
  - the raw GetData() impedance vector in the exact SDK buffer order
  - raw Ohm values and converted kOhm values
  - both zero-based and one-based column/index numbers
  - two enum interpretations:
      1. the C wrapper enum from wrapper.h used by this Python app
      2. the older C++ enum from channel.h / API documentation
  - the Appendix-A CA-200/201 manual order as a reference table

The SDK manual says the impedance "stream" is actually a state. Calling GetData()
returns one latest-known impedance state and does not need to be polled once per
second. Values are returned in Ohm; the app displays kOhm by dividing by 1000.

Example:
    py -3.13 diagnose_impedance_channels.py --duration 20 --layout capfiles/waveguard64.txt

Outputs:
    impedance_diagnostic_summary.csv
    impedance_diagnostic_long.csv
    impedance_diagnostic_wide.csv
    impedance_diagnostic.xlsx   (if openpyxl is installed)
"""
from __future__ import annotations

import argparse
import csv
import math
import statistics
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from eego_sdk import EegoSdk, Channel
from layout import load_layout, reference_electrodes


# C wrapper enum as defined in sdk/include/eemagine/sdk/wrapper.h
# This is what our ctypes wrapper receives from eemagine_sdk_channel_info.type.
WRAPPER_ENUM = {
    0: "REFERENCE",
    1: "BIPOLAR",
    2: "ACCELEROMETER",
    3: "GYROSCOPE",
    4: "MAGNETOMETER",
    5: "TRIGGER",
    6: "SAMPLE_COUNTER",
    7: "IMPEDANCE_REFERENCE",
    8: "IMPEDANCE_GROUND",
}

# Older/high-level C++ SDK enum from channel.h / API description.
# This is included only for comparison because old C++ examples often use these
# numeric codes directly: reference=1, impedance_reference=5, impedance_ground=6.
CPP_ENUM = {
    0: "none",
    1: "reference",
    2: "bipolar",
    3: "trigger",
    4: "sample_counter",
    5: "impedance_reference",
    6: "impedance_ground",
    7: "accelerometer",
    8: "gyroscope",
    9: "magnetometer",
}


@dataclass(frozen=True)
class ManualContact:
    manual_position: int
    ref_number: int | None
    electrode: str
    connector: str
    connector_row_label: str
    active_pin: int | None
    shield_pin: int | None
    role: str


# Appendix A order for CA-200/201, eego amplifier EE-21x/EE-22x.
# The important fact is that EOG is Ref 32. REF and GND are separate contacts.
MANUAL_REF_ORDER = [
    "Fp1", "Fpz", "Fp2", "F7", "F3", "Fz", "F4", "F8",
    "FC5", "FC1", "FC2", "FC6", "M1", "T7", "C3", "Cz",
    "C4", "T8", "M2", "CP5", "CP1", "CP2", "CP6", "P7",
    "P3", "Pz", "P4", "P8", "POz", "O1", "O2", "EOG",
    "AF7", "AF3", "AF4", "AF8", "F5", "F1", "F2", "F6",
    "FC3", "FCz", "FC4", "C5", "C1", "C2", "C6", "CP3",
    "CP4", "P5", "P1", "P2", "P6", "PO5", "PO3", "PO4",
    "PO6", "FT7", "FT8", "TP7", "TP8", "PO7", "PO8", "Oz",
]


def build_manual_contacts() -> list[ManualContact]:
    contacts: list[ManualContact] = []
    # EEG IN 1: GND, Ref 1-32, REF
    contacts.append(ManualContact(0, None, "GND", "EEG IN 1", "GND", 34, 68, "ground"))
    for i, name in enumerate(MANUAL_REF_ORDER[:32], start=1):
        # active pins Ref1..Ref32 = 33..2; shield = 67..36
        contacts.append(ManualContact(len(contacts), i, name, "EEG IN 1", f"Ref {i}", 34 - i, 68 - i, "reference"))
    contacts.append(ManualContact(len(contacts), None, "REF", "EEG IN 1", "REF", 1, 35, "reference_aux"))
    # EEG IN 2: not connected, Ref 33-64, not connected
    contacts.append(ManualContact(len(contacts), None, "not connected", "EEG IN 2", "not connected", 34, 68, "not_connected"))
    for i, name in enumerate(MANUAL_REF_ORDER[32:], start=33):
        # active pins Ref33..Ref64 = 33..2; shield = 67..36
        local = i - 32
        contacts.append(ManualContact(len(contacts), i, name, "EEG IN 2", f"Ref {i}", 34 - local, 68 - local, "reference"))
    contacts.append(ManualContact(len(contacts), None, "not connected", "EEG IN 2", "not connected", 1, 35, "not_connected"))
    return contacts


MANUAL_CONTACTS = build_manual_contacts()
MANUAL_REF_BY_INDEX0 = {i: name for i, name in enumerate(MANUAL_REF_ORDER)}
MANUAL_REF_BY_REFNO = {i + 1: name for i, name in enumerate(MANUAL_REF_ORDER)}


def ref_mask(n: int) -> int:
    n = max(0, min(64, n))
    if n >= 64:
        return 0xFFFFFFFFFFFFFFFF
    if n == 0:
        return 0
    return (1 << n) - 1


def fmt_float(x: float | None, digits: int = 6) -> str:
    if x is None or (isinstance(x, float) and (math.isnan(x) or math.isinf(x))):
        return ""
    return f"{float(x):.{digits}f}"


def wrapper_type(code: int) -> str:
    return WRAPPER_ENUM.get(int(code), f"UNKNOWN_WRAPPER_{code}")


def cpp_type(code: int) -> str:
    return CPP_ENUM.get(int(code), f"unknown_cpp_{code}")


def likely_impedance_kind(ch: Channel) -> str:
    """Classify by the C wrapper enum first; include old C++ code fallback.

    This is not used to hide any columns. It only adds a human-readable guess.
    Every SDK column is still written to the output.
    """
    if ch.type_code == 7 or ch.type.upper() == "IMPEDANCE_REFERENCE":
        return "impedance_reference_wrapper"
    if ch.type_code == 8 or ch.type.upper() == "IMPEDANCE_GROUND":
        return "impedance_ground_wrapper"
    # fallback for older C++ getType() numbers, only as a warning/comparison
    if ch.type_code == 5:
        return "could_be_cpp_impedance_reference_OR_wrapper_trigger"
    if ch.type_code == 6:
        return "could_be_cpp_impedance_ground_OR_wrapper_sample_counter"
    return "other"


def expected_from_manual_by_sdk_index(ch: Channel) -> dict[str, str | int | None]:
    """Map only from SDK-reported index to Appendix-A reference order.

    This does not use the buffer column number. It asks: if this SDK channel has
    index 0..63, which CA-200 electrode would that be according to Appendix A?
    """
    idx = int(ch.index)
    name = MANUAL_REF_BY_INDEX0.get(idx)
    if name is None:
        return {"manual_by_sdk_index": "", "manual_ref_number_by_sdk_index": ""}
    return {"manual_by_sdk_index": name, "manual_ref_number_by_sdk_index": idx + 1}


def expected_from_manual_by_impedance_ordinal(ordinal: int) -> dict[str, str | int | None]:
    """Map by ordinal among impedance-reference columns only.

    This asks: if we count only columns whose type looks like impedance_reference,
    which CA-200 electrode would this ordinal correspond to?
    """
    if 0 <= ordinal < 64:
        return {"manual_by_impedance_ordinal": MANUAL_REF_ORDER[ordinal], "manual_ref_number_by_impedance_ordinal": ordinal + 1}
    if ordinal == 64:
        return {"manual_by_impedance_ordinal": "REF? (extra impedance_reference after Ref64)", "manual_ref_number_by_impedance_ordinal": ""}
    if ordinal >= 65:
        return {"manual_by_impedance_ordinal": f"extra_impedance_reference_{ordinal - 63}", "manual_ref_number_by_impedance_ordinal": ""}
    return {"manual_by_impedance_ordinal": "", "manual_ref_number_by_impedance_ordinal": ""}


def channel_rows(channels: Iterable[Channel], source: str) -> list[dict]:
    rows = []
    for pos, ch in enumerate(channels):
        row = {
            "source": source,
            "position_0based": pos,
            "position_1based": pos + 1,
            "sdk_index_0based": int(ch.index),
            "sdk_index_1based": int(ch.index) + 1,
            "raw_type_code": int(ch.type_code),
            "python_wrapper_type": ch.type,
            "wrapper_h_type_interpretation": wrapper_type(ch.type_code),
            "cpp_channel_h_type_interpretation": cpp_type(ch.type_code),
            "likely_impedance_kind": likely_impedance_kind(ch),
        }
        row.update(expected_from_manual_by_sdk_index(ch))
        rows.append(row)
    return rows


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: dict[str, tuple[list[dict], list[str]]]) -> bool:
    """Write an .xlsx workbook if openpyxl is available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.formatting.rule import CellIsRule
    except Exception as exc:
        print(f"openpyxl not available; skipped XLSX output ({exc}). CSV files were still written.")
        return False

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    for sheet_name, (rows, fields) in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # widths
        for col_idx, field in enumerate(fields, start=1):
            width = max(len(str(field)) + 2, 10)
            for row_idx in range(2, min(ws.max_row, 80) + 1):
                width = max(width, min(len(str(ws.cell(row_idx, col_idx).value or "")) + 2, 42))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 42)
        # Add table if not too wide/empty
        if ws.max_row >= 2 and ws.max_column >= 1:
            table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            safe_name = "T_" + "".join(ch if ch.isalnum() else "_" for ch in sheet_name)[:25]
            tab = Table(displayName=safe_name, ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)
        # Conditional formatting for kOhm-ish columns.
        for col_idx, field in enumerate(fields, start=1):
            if "kOhm" in field or "kohm" in field.lower():
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["10"], fill=green_fill))
                ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["10", "20"], fill=yellow_fill))
                ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["20"], fill=red_fill))
    wb.save(path)
    return True


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--layout", default="capfiles/waveguard64.txt", help="Electrode layout file")
    parser.add_argument("--duration", type=float, default=15.0, help="Diagnostic duration in seconds")
    parser.add_argument("--poll", type=float, default=0.50, help="Sleep between impedance-state reads in seconds")
    parser.add_argument("--out", default="impedance_diagnostic", help="Output prefix")
    parser.add_argument("--amp", type=int, default=None, help="Amplifier id; default = first detected amplifier")
    parser.add_argument("--full-mask", action="store_true", help="Request all 64 referential impedance channels regardless of layout length")
    args = parser.parse_args()

    layout_path = Path(args.layout)
    electrodes = load_layout(layout_path)
    regular_names = [e.name for e in reference_electrodes(electrodes)][:64]
    requested_ref_count = 64 if args.full_mask else len(regular_names)
    mask = ref_mask(requested_ref_count)

    readme_rows = [
        {"field": "Purpose", "value": "Dump exactly what the eego SDK returns for impedance channel list and GetData buffer."},
        {"field": "Scale", "value": "SDK impedance GetData values are raw Ohm; kOhm = raw_ohm / 1000."},
        {"field": "Indexing", "value": "Every channel includes both zero-based and one-based column/index numbers."},
        {"field": "Enum warning", "value": "wrapper.h C enum differs numerically from the high-level C++ channel.h enum. Both interpretations are shown."},
        {"field": "Manual order", "value": "Appendix A CA-200/201 order: Ref 1 Fp1 ... Ref 32 EOG ... Ref 64 Oz; REF and GND are separate contacts."},
        {"field": "Requested ref_mask", "value": f"0x{mask:016X}"},
        {"field": "Layout", "value": str(layout_path)},
    ]

    with EegoSdk() as sdk:
        amps = sdk.amplifiers()
        if not amps:
            raise SystemExit("No eego amplifier detected.")
        amp = next((a for a in amps if a.id == args.amp), amps[0])
        sdk_version = sdk.version()
        print(f"Using amplifier serial={amp.serial!r}, id={amp.id}, sdk_version={sdk_version}")
        print(f"Layout: {layout_path} | layout reference contacts={len(regular_names)} | requested_ref_count={requested_ref_count} | ref_mask=0x{mask:016X}")

        amplifier_channels = sdk.amplifier_channels(amp.id)
        amplifier_rows = channel_rows(amplifier_channels, "amplifier.getChannelList_before_impedance")

        stream_id = sdk.open_impedance_stream(amp.id, ref_mask=mask)
        try:
            stream_channels = sdk.stream_channels(stream_id)
            stream_rows_base = channel_rows(stream_channels, "impedanceStream.getChannelList")

            print("\nExact SDK impedance stream channel list:")
            imp_ord = 0
            col_meta = []
            for col, ch in enumerate(stream_channels):
                is_imp_ref = likely_impedance_kind(ch).startswith("impedance_reference_wrapper") or ch.type_code == 5
                ordinal = imp_ord if is_imp_ref else -1
                if is_imp_ref:
                    imp_ord += 1
                by_ord = expected_from_manual_by_impedance_ordinal(ordinal)
                by_idx = expected_from_manual_by_sdk_index(ch)
                print(
                    f"  buffer_col0={col:03d} buffer_col1={col+1:03d} "
                    f"raw_type_code={ch.type_code:02d} wrapper={wrapper_type(ch.type_code):22s} "
                    f"cpp={cpp_type(ch.type_code):22s} sdk_index0={ch.index:03d} "
                    f"manual_by_index={by_idx.get('manual_by_sdk_index','')} "
                    f"manual_by_ordinal={by_ord.get('manual_by_impedance_ordinal','')}"
                )
                col_meta.append((col, ch, ordinal, by_idx, by_ord))

            latest_by_col: dict[int, list[float]] = {col: [] for col, *_ in col_meta}
            raw_latest_by_col: dict[int, list[float]] = {col: [] for col, *_ in col_meta}
            long_rows: list[dict] = []
            wide_rows: list[dict] = []
            deadline = time.time() + max(0.1, args.duration)
            state_count = 0
            print(f"\nCollecting impedance states for {args.duration:g} s; poll={args.poll:g} s...")
            while time.time() < deadline:
                sample = sdk.get_impedance_state(stream_id)
                t = time.time()
                wide = {"state_number": state_count, "time_unix": f"{t:.6f}", "returned_value_count": len(sample)}
                for col, ch, ordinal, by_idx, by_ord in col_meta:
                    if col >= len(sample):
                        continue
                    raw_ohm = float(sample[col])
                    kohm = raw_ohm / 1000.0
                    latest_by_col[col].append(kohm)
                    raw_latest_by_col[col].append(raw_ohm)
                    field_prefix = f"col{col:03d}_1b{col+1:03d}"
                    wide[f"{field_prefix}_raw_ohm"] = raw_ohm
                    wide[f"{field_prefix}_kOhm"] = kohm
                    row = {
                        "state_number": state_count,
                        "time_unix": f"{t:.6f}",
                        "buffer_column_0based": col,
                        "buffer_column_1based": col + 1,
                        "sdk_index_0based": int(ch.index),
                        "sdk_index_1based": int(ch.index) + 1,
                        "raw_type_code": int(ch.type_code),
                        "python_wrapper_type": ch.type,
                        "wrapper_h_type_interpretation": wrapper_type(ch.type_code),
                        "cpp_channel_h_type_interpretation": cpp_type(ch.type_code),
                        "likely_impedance_kind": likely_impedance_kind(ch),
                        "impedance_reference_ordinal_0based": ordinal if ordinal >= 0 else "",
                        "impedance_reference_ordinal_1based": ordinal + 1 if ordinal >= 0 else "",
                        "manual_by_sdk_index": by_idx.get("manual_by_sdk_index", ""),
                        "manual_ref_number_by_sdk_index": by_idx.get("manual_ref_number_by_sdk_index", ""),
                        "manual_by_impedance_ordinal": by_ord.get("manual_by_impedance_ordinal", ""),
                        "manual_ref_number_by_impedance_ordinal": by_ord.get("manual_ref_number_by_impedance_ordinal", ""),
                        "raw_ohm_exact": raw_ohm,
                        "kOhm_exact": kohm,
                    }
                    long_rows.append(row)
                wide_rows.append(wide)
                state_count += 1
                time.sleep(max(0.01, args.poll))
        finally:
            sdk.close_stream(stream_id)

    # Summary from the exact columns returned by the impedance stream.
    stream_summary_rows: list[dict] = []
    for col, ch, ordinal, by_idx, by_ord in col_meta:
        vals = latest_by_col.get(col, [])
        raw_vals = raw_latest_by_col.get(col, [])
        median_kohm = statistics.median(vals) if vals else None
        min_kohm = min(vals) if vals else None
        max_kohm = max(vals) if vals else None
        stdev_kohm = statistics.pstdev(vals) if len(vals) > 1 else 0.0 if vals else None
        row = {
            "buffer_column_0based": col,
            "buffer_column_1based": col + 1,
            "sdk_index_0based": int(ch.index),
            "sdk_index_1based": int(ch.index) + 1,
            "raw_type_code": int(ch.type_code),
            "python_wrapper_type": ch.type,
            "wrapper_h_type_interpretation": wrapper_type(ch.type_code),
            "cpp_channel_h_type_interpretation": cpp_type(ch.type_code),
            "likely_impedance_kind": likely_impedance_kind(ch),
            "impedance_reference_ordinal_0based": ordinal if ordinal >= 0 else "",
            "impedance_reference_ordinal_1based": ordinal + 1 if ordinal >= 0 else "",
            "manual_by_sdk_index": by_idx.get("manual_by_sdk_index", ""),
            "manual_ref_number_by_sdk_index": by_idx.get("manual_ref_number_by_sdk_index", ""),
            "manual_by_impedance_ordinal": by_ord.get("manual_by_impedance_ordinal", ""),
            "manual_ref_number_by_impedance_ordinal": by_ord.get("manual_ref_number_by_impedance_ordinal", ""),
            "n_states": len(vals),
            "median_raw_ohm": statistics.median(raw_vals) if raw_vals else "",
            "median_kOhm": median_kohm if median_kohm is not None else "",
            "min_kOhm": min_kohm if min_kohm is not None else "",
            "max_kOhm": max_kohm if max_kohm is not None else "",
            "stdev_kOhm": stdev_kohm if stdev_kohm is not None else "",
            "low_under_10kOhm": bool(median_kohm is not None and median_kohm < 10.0),
            "yellow_10_to_20kOhm": bool(median_kohm is not None and 10.0 <= median_kohm <= 20.0),
            "high_over_20kOhm": bool(median_kohm is not None and median_kohm > 20.0),
        }
        stream_summary_rows.append(row)

    low_sorted_rows = sorted(stream_summary_rows, key=lambda r: (float("inf") if r["median_kOhm"] == "" else float(r["median_kOhm"])))

    manual_rows = [
        {
            "manual_position": c.manual_position,
            "ref_number": c.ref_number if c.ref_number is not None else "",
            "electrode": c.electrode,
            "connector": c.connector,
            "connector_row_label": c.connector_row_label,
            "active_pin": c.active_pin if c.active_pin is not None else "",
            "shield_pin": c.shield_pin if c.shield_pin is not None else "",
            "role": c.role,
            "reference_index_0based_if_ref": (c.ref_number - 1) if c.ref_number is not None else "",
            "reference_index_1based_if_ref": c.ref_number if c.ref_number is not None else "",
        }
        for c in MANUAL_CONTACTS
    ]

    prefix = Path(args.out)
    paths = {
        "summary": prefix.with_name(prefix.name + "_summary.csv"),
        "long": prefix.with_name(prefix.name + "_long.csv"),
        "wide": prefix.with_name(prefix.name + "_wide.csv"),
        "amplifier_channels": prefix.with_name(prefix.name + "_amplifier_channels.csv"),
        "manual_order": prefix.with_name(prefix.name + "_manual_order.csv"),
        "xlsx": prefix.with_suffix(".xlsx"),
    }

    summary_fields = list(stream_summary_rows[0].keys()) if stream_summary_rows else []
    long_fields = list(long_rows[0].keys()) if long_rows else []
    wide_fields = list(wide_rows[0].keys()) if wide_rows else []
    amp_fields = list(amplifier_rows[0].keys()) if amplifier_rows else []
    manual_fields = list(manual_rows[0].keys()) if manual_rows else []
    readme_fields = ["field", "value"]

    write_csv(paths["summary"], stream_summary_rows, summary_fields)
    write_csv(paths["long"], long_rows, long_fields)
    write_csv(paths["wide"], wide_rows, wide_fields)
    write_csv(paths["amplifier_channels"], amplifier_rows, amp_fields)
    write_csv(paths["manual_order"], manual_rows, manual_fields)

    xlsx_written = write_xlsx(paths["xlsx"], {
        "README": (readme_rows, readme_fields),
        "Impedance_Summary": (stream_summary_rows, summary_fields),
        "Low_kOhm_Sorted": (low_sorted_rows, summary_fields),
        "Impedance_Long": (long_rows, long_fields),
        "Raw_State_Wide": (wide_rows, wide_fields),
        "Stream_ChannelList": (stream_rows_base, list(stream_rows_base[0].keys()) if stream_rows_base else []),
        "Amplifier_ChannelList": (amplifier_rows, amp_fields),
        "Manual_Appendix_A": (manual_rows, manual_fields),
    })

    print("\nSaved diagnostic files:")
    for key, path in paths.items():
        if key == "xlsx" and not xlsx_written:
            continue
        print(f"  {path}")

    print("\nHow to inspect:")
    print("  1. Open impedance_diagnostic.xlsx, sheet 'Impedance_Summary'.")
    print("  2. Sort by median_kOhm or open 'Low_kOhm_Sorted'.")
    print("  3. Compare buffer_column_1based and sdk_index_1based against Manual_Appendix_A.")
    print("  4. Do not trust only one mapping column; check both manual_by_sdk_index and manual_by_impedance_ordinal.")
    print("  5. The raw buffer order is preserved in Raw_State_Wide and Impedance_Long.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
